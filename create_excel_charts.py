"""
Create Excel Charts for Heckscher-Ohlin Analysis
This script generates an Excel file with the same charts as the Python analysis:
1. Dual-axis chart: Capital Deepening vs Capital-Labor Ratio
2. Scatter plot: Real Exports vs Capital-Labor Ratio (with trendline)
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, ScatterChart, Reference, Series
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.trendline import Trendline
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os


def create_excel_with_charts():
    """Create Excel file with data and charts"""

    # Load the data
    csv_path = "heckscher_ohlin_data.csv"
    if not os.path.exists(csv_path):
        print(
            f"Error: {csv_path} not found. Please run heckscher_ohlin_analysis.py first."
        )
        return

    df = pd.read_csv(csv_path)
    print(f"Loaded data with {len(df)} rows")

    # Create workbook
    wb = Workbook()

    # =========================================================================
    # Sheet 1: Data
    # =========================================================================
    ws_data = wb.active
    ws_data.title = "Data"

    # Write headers with formatting
    headers = [
        "Year",
        "Real_GDP",
        "Labor_Force",
        "Real_Investment",
        "Real_Exports",
        "Capital_Deepening_Pct",
        "Capital_Labor_Ratio",
    ]

    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Write data
    for row_idx, row in df.iterrows():
        for col_idx, header in enumerate(headers):
            cell = ws_data.cell(row=row_idx + 2, column=col_idx + 1, value=row[header])
            cell.border = thin_border
            if col_idx >= 5:  # Format percentages and ratios
                cell.number_format = "#,##0.00"
            elif col_idx >= 1 and col_idx <= 4:
                cell.number_format = "#,##0.00"

    # Adjust column widths
    column_widths = [8, 15, 15, 18, 15, 22, 20]
    for col, width in enumerate(column_widths, 1):
        ws_data.column_dimensions[chr(64 + col)].width = width

    # =========================================================================
    # Sheet 2: Dual-Axis Chart (Capital Deepening & K/L Ratio)
    # As per instructions: Left Axis = Capital Deepening (%), Right Axis = K/L Ratio ($)
    # =========================================================================
    ws_chart1 = wb.create_sheet("Dual-Axis Chart")

    # Headers with formatting
    ws_chart1["A1"] = "Year"
    ws_chart1["B1"] = "Capital Deepening (%)"
    ws_chart1["C1"] = "Capital-Labor Ratio ($)"

    for col in range(1, 4):
        cell = ws_chart1.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for i, row in enumerate(df.iterrows(), 2):
        idx, data = row
        # Store year as STRING so it displays as category label on x-axis
        ws_chart1.cell(row=i, column=1, value=str(int(data["Year"]))).border = (
            thin_border
        )
        ws_chart1.cell(row=i, column=2, value=data["Capital_Deepening_Pct"]).border = (
            thin_border
        )
        ws_chart1.cell(row=i, column=2).number_format = "#,##0.00"
        ws_chart1.cell(row=i, column=3, value=data["Capital_Labor_Ratio"]).border = (
            thin_border
        )
        ws_chart1.cell(row=i, column=3).number_format = "#,##0.00"

    # Adjust column widths
    ws_chart1.column_dimensions["A"].width = 8
    ws_chart1.column_dimensions["B"].width = 22
    ws_chart1.column_dimensions["C"].width = 22

    num_rows = len(df) + 1

    # Create PRIMARY line chart for Capital Deepening (LEFT AXIS)
    chart1 = LineChart()
    chart1.title = "U.S. Capital Deepening and Capital-Labor Ratio (1960-Present)"
    chart1.style = 10
    chart1.width = 20
    chart1.height = 12
    chart1.y_axis.title = "Capital Deepening (%)"
    chart1.x_axis.title = "Year"

    # X-axis category labels settings
    chart1.x_axis.tickLblSkip = 5  # Show every 5th label
    chart1.x_axis.tickMarkSkip = 5

    # Add Capital Deepening data (primary/left axis)
    data1 = Reference(ws_chart1, min_col=2, min_row=1, max_row=num_rows)
    # Categories are the years in column A (row 2 to end, as row 1 is header)
    cats = Reference(ws_chart1, min_col=1, min_row=2, max_row=num_rows)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats)

    # Style the first series (blue line) - Capital Deepening
    s1 = chart1.series[0]
    s1.graphicalProperties.line.solidFill = "1F77B4"
    s1.graphicalProperties.line.width = 25000
    s1.marker.symbol = "circle"
    s1.marker.size = 4
    s1.marker.graphicalProperties.solidFill = "1F77B4"
    s1.marker.graphicalProperties.line.solidFill = "1F77B4"

    # Create SECONDARY line chart for K/L Ratio (RIGHT AXIS)
    chart2 = LineChart()
    chart2.y_axis.axId = 200
    chart2.y_axis.title = "K/L Ratio ($)"

    # Add K/L Ratio data (secondary/right axis)
    data2 = Reference(ws_chart1, min_col=3, min_row=1, max_row=num_rows)
    chart2.add_data(data2, titles_from_data=True)

    # Style the second series (red line) - K/L Ratio
    s2 = chart2.series[0]
    s2.graphicalProperties.line.solidFill = "D62728"
    s2.graphicalProperties.line.width = 25000
    s2.marker.symbol = "square"
    s2.marker.size = 4
    s2.marker.graphicalProperties.solidFill = "D62728"
    s2.marker.graphicalProperties.line.solidFill = "D62728"

    # Combine charts - secondary axis on right side
    chart1.y_axis.crosses = "min"
    chart2.y_axis.crosses = "max"
    chart1 += chart2

    ws_chart1.add_chart(chart1, "E2")

    # =========================================================================
    # Sheet 3: Regression Scatter Plot
    # =========================================================================
    ws_chart2 = wb.create_sheet("Regression Plot")

    # Copy Capital_Labor_Ratio and Real_Exports for the scatter plot
    ws_chart2["A1"] = "Capital-Labor Ratio ($)"
    ws_chart2["B1"] = "Real Exports (Billions $)"

    for i, row in enumerate(df.iterrows(), 2):
        idx, data = row
        ws_chart2.cell(row=i, column=1, value=data["Capital_Labor_Ratio"])
        ws_chart2.cell(row=i, column=2, value=data["Real_Exports"])

    # Create scatter chart
    scatter_chart = ScatterChart()
    scatter_chart.title = "Regression: Real Exports vs Capital-Labor Ratio"
    scatter_chart.style = 10
    scatter_chart.width = 18
    scatter_chart.height = 12
    scatter_chart.x_axis.title = "Capital-Labor Ratio ($ per Worker)"
    scatter_chart.y_axis.title = "Real Exports (Billions of Chained $)"

    # Add data
    xvalues = Reference(ws_chart2, min_col=1, min_row=2, max_row=num_rows)
    yvalues = Reference(ws_chart2, min_col=2, min_row=2, max_row=num_rows)
    series = Series(yvalues, xvalues, title="Observed Data")
    scatter_chart.series.append(series)

    # Add linear trendline
    series.trendline = Trendline(trendlineType="linear", dispRSqr=True, dispEq=True)

    # Style the scatter points
    series.graphicalProperties.line.noFill = True
    series.marker.symbol = "circle"
    series.marker.size = 7
    series.marker.graphicalProperties.solidFill = "1F77B4"

    ws_chart2.add_chart(scatter_chart, "D2")

    # =========================================================================
    # Sheet 4: Summary Statistics
    # =========================================================================
    ws_summary = wb.create_sheet("Summary")

    # Add summary statistics
    ws_summary["A1"] = "Summary Statistics"
    ws_summary["A1"].font = Font(bold=True, size=14)

    stats = [
        ("Variable", "Mean", "Std Dev", "Min", "Max"),
        (
            "Capital Deepening (%)",
            df["Capital_Deepening_Pct"].mean(),
            df["Capital_Deepening_Pct"].std(),
            df["Capital_Deepening_Pct"].min(),
            df["Capital_Deepening_Pct"].max(),
        ),
        (
            "Capital-Labor Ratio ($)",
            df["Capital_Labor_Ratio"].mean(),
            df["Capital_Labor_Ratio"].std(),
            df["Capital_Labor_Ratio"].min(),
            df["Capital_Labor_Ratio"].max(),
        ),
        (
            "Real Exports (Billions $)",
            df["Real_Exports"].mean(),
            df["Real_Exports"].std(),
            df["Real_Exports"].min(),
            df["Real_Exports"].max(),
        ),
        (
            "Real GDP (Billions $)",
            df["Real_GDP"].mean(),
            df["Real_GDP"].std(),
            df["Real_GDP"].min(),
            df["Real_GDP"].max(),
        ),
    ]

    for row_idx, row_data in enumerate(stats, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 3:
                cell.fill = header_fill
                cell.font = header_font
            elif col_idx > 1:
                cell.number_format = "#,##0.00"

    # Adjust column widths
    ws_summary.column_dimensions["A"].width = 25
    for col in ["B", "C", "D", "E"]:
        ws_summary.column_dimensions[col].width = 15

    # =========================================================================
    # Save the workbook
    # =========================================================================
    output_file = "heckscher_ohlin_charts.xlsx"
    wb.save(output_file)
    print(f"\n✓ Excel file saved: {output_file}")
    print("\nSheets created:")
    print("  1. Data - Complete dataset")
    print("  2. Dual-Axis Chart - Capital Deepening vs K/L Ratio")
    print("  3. Regression Plot - Real Exports vs K/L Ratio with trendline")
    print("  4. Summary - Summary statistics")

    return output_file


if __name__ == "__main__":
    create_excel_with_charts()
